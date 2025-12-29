import csv
from typing import List, Dict
import pandas as pd
from studytracker.db import Database

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.formatting.rule import CellIsRule
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ReportGenerator:
    
    def __init__(self, db: Database):
        self.db = db
    
    def export_courses_to_csv(self, filename: str):
        query = "SELECT id, name, teacher, credits FROM courses ORDER BY name"
        rows = self.db.fetch_all(query)
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['ID', 'Course Name', 'Teacher', 'Credits'])
            for row in rows:
                writer.writerow([row['id'], row['name'], row['teacher'], row['credits']])
        
        print(f"Courses exported to {filename}")
    
    def export_assignments_to_csv(self, filename: str):
        query = """
            SELECT 
                a.id,
                c.name as course_name,
                a.title,
                a.due_date,
                a.grade
            FROM assignments a
            JOIN courses c ON a.course_id = c.id
            ORDER BY a.due_date
        """
        rows = self.db.fetch_all(query)
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['ID', 'Course', 'Assignment', 'Due Date', 'Grade'])
            for row in rows:
                grade = row['grade'] if row['grade'] is not None else 'Not graded'
                writer.writerow([row['id'], row['course_name'], row['title'], 
                               row['due_date'], grade])
        
        print(f"Assignments exported to {filename}")
    
    def export_full_report_to_csv(self, filename: str):
        query = """
            SELECT 
                c.name as course_name,
                c.teacher,
                c.credits,
                a.title as assignment_title,
                a.due_date,
                a.grade
            FROM courses c
            LEFT JOIN assignments a ON c.id = a.course_id
            ORDER BY c.name, a.due_date
        """
        rows = self.db.fetch_all(query)
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['Course', 'Teacher', 'Credits', 'Assignment', 'Due Date', 'Grade'])
            for row in rows:
                grade = row['grade'] if row['grade'] is not None else 'Not graded'
                assignment = row['assignment_title'] if row['assignment_title'] else 'No assignments'
                due_date = row['due_date'] if row['due_date'] else ''
                writer.writerow([row['course_name'], row['teacher'], row['credits'],
                               assignment, due_date, grade])
        
        print(f"Full report exported to {filename}")
    
    def export_courses_to_excel(self, filename: str):
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl library is not installed. Run: pip install openpyxl")
            return
        
        query = "SELECT id, name, teacher, credits FROM courses ORDER BY name"
        rows = self.db.fetch_all(query)
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Courses"
        
        ws.append(['ID', 'Course Name', 'Teacher', 'Credits'])
        for row in rows:
            ws.append([row['id'], row['name'], row['teacher'], row['credits']])
        
        wb.save(filename)
        print(f"Courses exported to {filename}")
    
    def export_assignments_to_excel(self, filename: str):
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl library is not installed. Run: pip install openpyxl")
            return
        
        query = """
            SELECT 
                a.id,
                c.name as course_name,
                a.title,
                a.due_date,
                a.grade
            FROM assignments a
            JOIN courses c ON a.course_id = c.id
            ORDER BY a.due_date
        """
        rows = self.db.fetch_all(query)
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Assignments"
        
        # Write header
        ws.append(['ID', 'Course', 'Assignment', 'Due Date', 'Grade'])
        
        # Write data
        for row in rows:
            grade = row['grade'] if row['grade'] is not None else 'Not graded'
            ws.append([row['id'], row['course_name'], row['title'], 
                      row['due_date'], grade])
        
        # Save the workbook
        self._apply_grade_conditional_formatting(ws, grade_col="E", last_row=ws.max_row)
        wb.save(filename)
        print(f"Assignments exported to {filename}")
    
    def export_full_report_to_excel(self, filename: str):
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl library is not installed. Run: pip install openpyxl")
            return
        
        # Create workbook
        wb = Workbook()
        
        # Sheet 1: Courses
        ws_courses = wb.active
        ws_courses.title = "Courses"
        courses_query = "SELECT id, name, teacher, credits FROM courses ORDER BY name"
        courses = self.db.fetch_all(courses_query)
        ws_courses.append(['ID', 'Course Name', 'Teacher', 'Credits'])
        for row in courses:
            ws_courses.append([row['id'], row['name'], row['teacher'], row['credits']])
        
        # Sheet 2: Assignments
        ws_assignments = wb.create_sheet("Assignments")
        assignments_query = """
            SELECT 
                a.id,
                c.name as course_name,
                a.title,
                a.due_date,
                a.grade
            FROM assignments a
            JOIN courses c ON a.course_id = c.id
            ORDER BY a.due_date
        """
        assignments = self.db.fetch_all(assignments_query)
        ws_assignments.append(['ID', 'Course', 'Assignment', 'Due Date', 'Grade'])
        for row in assignments:
            grade = row['grade'] if row['grade'] is not None else 'Not graded'
            ws_assignments.append([row['id'], row['course_name'], row['title'],
                                  row['due_date'], grade])
        
        # Save the workbook
        self._apply_grade_conditional_formatting(ws_assignments, grade_col="E", last_row=ws_assignments.max_row)
        wb.save(filename)
        print(f"Full report exported to {filename}")

    def calculate_weighted_final_grade(self) -> float:
        query = """
            SELECT a.grade, c.credits
            FROM assignments a
            JOIN courses c ON a.course_id = c.id
            WHERE a.grade IS NOT NULL
        """
        rows = self.db.fetch_all(query)
        if not rows:
            return 0.0

        total_weighted = 0.0
        total_credits = 0
        for row in rows:
            total_weighted += row["grade"] * row["credits"]
            total_credits += row["credits"]

        return round(total_weighted / total_credits, 2) if total_credits else 0.0

    def export_full_report_with_pandas(self, filename: str, file_format: str = "csv"):
        query = """
            SELECT 
                c.name as course_name,
                c.teacher,
                c.credits,
                a.title as assignment_title,
                a.due_date,
                a.grade
            FROM courses c
            LEFT JOIN assignments a ON c.id = a.course_id
            ORDER BY c.name, a.due_date
        """
        rows = self.db.fetch_all(query)
        if not rows:
            print("No data to export.")
            return

        df = pd.DataFrame(rows, columns=rows[0].keys())

        final_grade = self.calculate_weighted_final_grade()
        summary_row = {
            "course_name": "WEIGHTED_FINAL_GRADE",
            "teacher": "",
            "credits": "",
            "assignment_title": "weighted by course credits",
            "due_date": "",
            "grade": final_grade,
        }
        df = pd.concat([df, pd.DataFrame([summary_row])], ignore_index=True)

        if file_format == "excel":
            try:
                df.to_excel(filename, index=False)
                if EXCEL_AVAILABLE:
                    wb = load_workbook(filename)
                    ws = wb.active
                    last_row = ws.max_row
                    self._apply_grade_conditional_formatting(ws, grade_col="F", last_row=last_row)

                    # Make the summary label stand out for quick scanning
                    summary_label_cell = ws[f"A{last_row}"]
                    summary_label_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    summary_label_cell.font = Font(bold=True)

                    wb.save(filename)
            except ImportError:
                print("Error: openpyxl is required for Excel export. Run: pip install openpyxl")
                return
        else:
            df.to_csv(filename, index=False)

        print(f"Pandas report exported to {filename}")

    def _apply_grade_conditional_formatting(self, ws, grade_col: str, last_row: int):
        if not EXCEL_AVAILABLE or last_row < 2:
            return

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        orange_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        red_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

        cell_range = f"{grade_col}2:{grade_col}{last_row}"

        ws.conditional_formatting.add(cell_range, CellIsRule(operator=">", formula=[70], fill=green_fill))
        ws.conditional_formatting.add(cell_range, CellIsRule(operator="between", formula=[50, 70], fill=orange_fill))
        ws.conditional_formatting.add(cell_range, CellIsRule(operator="<", formula=[50], fill=red_fill))
